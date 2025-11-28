# v20251023

import random
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from ortools.sat.python import cp_model

#### PART1 ####
# === 1.1 Setup & Data Loading ===
print("➡️  Solving...")

INPUT_ROSTER = "Roster_input.xlsx"
SHEET_NAME   = "Sheet1"
OUTPUT_FILE  = "Roster_Output1.xlsx"

wb_roster = load_workbook(INPUT_ROSTER)
ws_roster = wb_roster[SHEET_NAME]

# Month
month = ws_roster.cell(row=1, column=3).value


# === 1.2 Staff & quotas ===
ROW_START = 6
staff, ranks, quotas = [], [], {}

# open a second copy of the roster for formula results
wb_roster_values = load_workbook(INPUT_ROSTER, data_only=True)
ws_values        = wb_roster_values[SHEET_NAME]

r = ROW_START
while True:
    name = ws_roster.cell(row=r, column=4).value  # D
    if not name:
        break
    name = str(name).strip()
    staff.append(name)
    ranks.append(str(ws_roster.cell(row=r, column=3).value or "").strip())  # C

    # raw quotas E–O (static numbers, no formulas here)
    raw = [ws_roster.cell(row=r, column=c).value for c in range(5, 15)]
    quotas[name] = {
        "N*":      int(raw[0] or 0),
        "N":        int(raw[1] or 0),
        "N3":       int(raw[2] or 0),
        "Z":        int(raw[3] or 0),
        "Nspacing": float(raw[4] or 4),
        "SUN-Off":  int(raw[5]) if raw[5] is not None else None,
        "WE-Off":   int(raw[6]) if raw[6] is not None else None,
        "SUN P":    int(raw[7]) if raw[7] is not None else None,
        "min_pa":   float(raw[8] or 0),
        "max_pa":   float(raw[9] or 1.5),
    }

    # fetch the formula result for init_hr from the values‐only sheet:
    init_val = ws_values.cell(row=r, column=15).value  # column O (15)
    quotas[name]["init_hr"] = float(init_val or 0)
    ws_roster.cell(row=r, column=15, value=init_val)

    r += 1

NUM_STAFF = len(staff)

# === Global constraints - Min SUN Off, Min WE Off, Max SUN P, Max P/A ratio
min_sun_off = int(ws_roster["J2"].value or 0)
min_we_off  = int(ws_roster["K2"].value or 0)
max_sun_pm = int(ws_roster["L3"].value or 99)

max_sun_off = int(ws_roster["J3"].value or 99)   # default high if blank
max_we_off  = int(ws_roster["K3"].value or 99)   # default high if blank
min_sun_pm  = int(ws_roster["L2"].value or 0)    # default 0 if blank


globalmax_pa = int(ws_roster["N2"].value or 1)
global_Nspacing = int(ws_roster["I2"].value or 4)


# 1.3 Locate Day Block
col = 1
while True:
    if str(ws_roster.cell(row=1, column=col).value or "").strip().upper() == "START":
        START_COL = col + 1
        break
    col += 1

# 1.4 Read days
special, dates, weekdays = [], [], []
col = START_COL
while True:
    tag = ws_roster.cell(row=1, column=col).value
    date = ws_roster.cell(row=3, column=col).value
    wd   = ws_roster.cell(row=4, column=col).value
    if date is None:
        break
    special.append(str(tag or "").upper().strip())
    dates.append(date.day if isinstance(date, datetime.datetime) else int(date))
    weekdays.append(str(wd or "").upper().strip())
    col += 1

NUM_DAYS = len(dates)
LAST_COL = START_COL + NUM_DAYS - 1


# === 1.5 Read per-day coverage targets from 'Manpower' block ===
# Find the row and column of the cell containing 'Manpower'
manpower_row, manpower_col = None, None
for row in range(1, 100):  # Search within first 100 rows
    for col in range(1, ws_roster.max_column + 1):
        val = str(ws_roster.cell(row=row, column=col).value or "").strip()
        if val.upper() == "MANPOWER":
            manpower_row, manpower_col = row, col
            break
    if manpower_row:
        break
if manpower_row is None:
    raise ValueError("❌ Could not locate 'Manpower' cell in Roster sheet.")

    
# 1.6 Classify days
day_type = []
for i in range(NUM_DAYS):
    tag, wd = special[i], weekdays[i]
    if wd in ("SAT", "SUN") or tag.startswith(("SH", "PH")):
        day_type.append("WE")   # Weekend / Holiday
    else:
        day_type.append("WD")   # Regular weekday

# 1.7 Duty requests
fixed_raw, fixed_clean = {}, {}
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS):
        cell = ws_roster.cell(row=ROW_START+s, column=START_COL+d)
        v = cell.value
        if isinstance(v, str) and v.strip():
            orig = v.strip()
            fixed_raw[(s,d)] = orig
            fixed_clean[(s,d)] = orig.rstrip("↗").strip()
            # fixed_clean[(s,d)] = orig.rstrip("↗").strip().upper()



#### PART 2 ####
# === 2. Model & Assignment ===
# 2.1 Coverage targets

# 2.2 Shift sets
SHIFTS   = ["A","P","N*","N","N3","O","Z","T","½t","½Z","AL","☆","½●½O","兒","父","SH","PH"]
AM_SH    = {"A"}
PM_SH    = {"P"}
NIGHT_SH = {"N*","N","N3"}
OFF      = {"O","AL","☆","½●½O","兒","父","SH","PH"}

model = cp_model.CpModel()
X = {}
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS):
        for t in SHIFTS:
            X[s,d,t] = model.NewBoolVar(f"x_{s}_{d}_{t}")



# Read coverage targets for each day by column
cov_targets_per_day = []  # List of tuples: (AM, PM, Night)
for offset in range(NUM_DAYS):
    col = START_COL + offset
    am = int(ws_roster.cell(row=manpower_row + 1, column=col).value or 0)
    pm = int(ws_roster.cell(row=manpower_row + 2, column=col).value or 0)
    nt = int(ws_roster.cell(row=manpower_row + 3, column=col).value or 0)
    cov_targets_per_day.append((am, pm, nt))

# === Read senior constraints (min/max for A and P) ===
minsenA_per_day, maxsenA_per_day = [], []
minsenP_per_day, maxsenP_per_day = [], []

for offset in range(NUM_DAYS):
    col = START_COL + offset
    # seniors (you already had these)
    minsenA = int(ws_roster.cell(row=manpower_row + 4, column=col).value or 0)
    maxsenA = int(ws_roster.cell(row=manpower_row + 5, column=col).value or 99)
    minsenP = int(ws_roster.cell(row=manpower_row + 6, column=col).value or 0)
    maxsenP = int(ws_roster.cell(row=manpower_row + 7, column=col).value or 99)
    minsenA_per_day.append(minsenA); maxsenA_per_day.append(maxsenA)
    minsenP_per_day.append(minsenP); maxsenP_per_day.append(maxsenP)

# === Cadre index sets ===
# Seniors: CON*, AC, HT1
senior_idx = {s for s, r in enumerate(ranks) if str(r).upper().startswith(("CON", "AC", "HT1"))}
# Juniors: HT2, BT, E
junior_idx = {s for s, r in enumerate(ranks) if str(r).upper().startswith(("HT2", "BT", "E"))}

# Extra groups you asked for
con_idx   = {s for s, r in enumerate(ranks) if str(r).upper().startswith("CON")}
ac_idx    = {s for s, r in enumerate(ranks) if str(r).upper().startswith("AC")}
conac_idx = con_idx | ac_idx
ht_idx    = {s for s, r in enumerate(ranks) if str(r).upper().startswith(("HT1","HT2"))}
bt_idx    = {s for s, r in enumerate(ranks) if str(r).upper().startswith("BT")}
e_idx     = {s for s, r in enumerate(ranks) if str(r).upper().startswith("E")}

# === Helpers to read & enforce per-day min/max A/P for any group ===
def _read_minmax_rows(r_minA, r_maxA, r_minP, r_maxP):
    minA, maxA, minP, maxP = [], [], [], []
    for off in range(NUM_DAYS):
        c = START_COL + off
        def _iv(v, dflt): 
            try: return int(v) if v not in (None, "") else dflt
            except: return dflt
        minA.append(_iv(ws_roster.cell(row=r_minA, column=c).value, 0))
        maxA.append(_iv(ws_roster.cell(row=r_maxA, column=c).value, 99))
        minP.append(_iv(ws_roster.cell(row=r_minP, column=c).value, 0))
        maxP.append(_iv(ws_roster.cell(row=r_maxP, column=c).value, 99))
    return minA, maxA, minP, maxP

def _enforce_group_minmax(group_idx, minA, maxA, minP, maxP):
    for d in range(NUM_DAYS):
        am_cov, pm_cov, _ = cov_targets_per_day[d]
        # clamp to coverage
        a_min = min(minA[d], am_cov); a_max = min(maxA[d], am_cov)
        p_min = min(minP[d], pm_cov); p_max = min(maxP[d], pm_cov)
        model.Add(sum(X[s, d, "A"] for s in group_idx) >= a_min)
        model.Add(sum(X[s, d, "A"] for s in group_idx) <= a_max)
        model.Add(sum(X[s, d, "P"] for s in group_idx) >= p_min)
        model.Add(sum(X[s, d, "P"] for s in group_idx) <= p_max)

# === Read per-day min/max rows for each extra group ===
# Adjust these offsets to match your sheet:
junA_min, junA_max, junP_min, junP_max = _read_minmax_rows(manpower_row+8,  manpower_row+9,  manpower_row+10, manpower_row+11)  # Juniors
caA_min,  caA_max,  caP_min,  caP_max  = _read_minmax_rows(manpower_row+12, manpower_row+13, manpower_row+14, manpower_row+15)  # CON+AC
acA_min,  acA_max,  acP_min,  acP_max  = _read_minmax_rows(manpower_row+16, manpower_row+17, manpower_row+18, manpower_row+19)  # AC only
htA_min,  htA_max,  htP_min,  htP_max  = _read_minmax_rows(manpower_row+20, manpower_row+21, manpower_row+22, manpower_row+23)  # HT (HT1+HT2)
btA_min,  btA_max,  btP_min,  btP_max  = _read_minmax_rows(manpower_row+24, manpower_row+25, manpower_row+26, manpower_row+27)  # BT
eA_min,   eA_max,   eP_min,   eP_max   = _read_minmax_rows(manpower_row+28, manpower_row+29, manpower_row+30, manpower_row+31)  # E

# === Enforce (keep your existing senior enforcement wherever it lives) ===
_enforce_group_minmax(junior_idx, junA_min, junA_max, junP_min, junP_max)
_enforce_group_minmax(conac_idx,   caA_min,  caA_max,  caP_min,  caP_max)
_enforce_group_minmax(ac_idx,      acA_min,  acA_max,  acP_min,  acP_max)
_enforce_group_minmax(ht_idx,      htA_min,  htA_max,  htP_min,  htP_max)
_enforce_group_minmax(bt_idx,      btA_min,  btA_max,  btP_min,  btP_max)
_enforce_group_minmax(e_idx,       eA_min,   eA_max,   eP_min,   eP_max)


# one shift/day
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS):
        model.AddExactlyOne(X[s,d,t] for t in SHIFTS)

# Forbid Z, T, ½t, etc unless explicitly pre-filled in the sheet
SPECIAL_FIXED_ONLY = {"Z", "T", "½t","½Z","AL","☆","½●½O","兒","父","SH","PH"}

for s in range(NUM_STAFF):
    for d in range(NUM_DAYS):
        for sp in SPECIAL_FIXED_ONLY:
            # only allow if user pre-filled exactly this code
            if not ((s, d) in fixed_clean and fixed_clean[(s, d)] == sp):
                model.Add(X[s, d, sp] == 0)


# === Read last 7 days of LAST month (reference only) — shifted one col left ===
from openpyxl.utils import get_column_letter

def split_arrow(val):
    """Return (core, has_arrow). core has any trailing ↗ removed."""
    s = str(val or "").strip()
    if s.endswith("↗"):
        return s[:-1].strip(), True
    return s, False

# safety: now we need 8 columns before START (to reach START_COL - 8)
if START_COL <= 8:
    raise RuntimeError("Not enough columns before START to read last 7 days (shifted left).")

# Columns for last week (left→right): START_COL-8 ... START_COL-2
last7_cols = list(range(START_COL - 8, START_COL - 1))
last7_letters = [get_column_letter(c) for c in last7_cols]

# Per-staff arrays
# prev_last7[s][i]  -> clean duty code at row s, col last7_cols[i]
# prev_last7_raw[s] -> raw cell value (keeps ↗ if present)
# prev_last7_arrow[s][i] -> True if that cell had ↗
prev_last7       = [[None]*7 for _ in range(NUM_STAFF)]
prev_last7_raw   = [[None]*7 for _ in range(NUM_STAFF)]
prev_last7_arrow = [[False]*7 for _ in range(NUM_STAFF)]

for s in range(NUM_STAFF):
    row = ROW_START + s
    for i, c in enumerate(last7_cols):
        raw = ws_roster.cell(row=row, column=c).value
        core, has_arr = split_arrow(raw)
        if core == "":
            core = "O"  # empty -> Off
        prev_last7[s][i]       = core
        prev_last7_raw[s][i]   = raw
        prev_last7_arrow[s][i] = has_arr

# === Enforce duty‐request overrides ===
for (s, d), clean in fixed_clean.items():
    orig = fixed_raw[(s, d)]
    arrow = orig.endswith("↗")  # (unused here, ok to keep/remove)
        
    # map request to allowed shifts
    if clean == "A":
        allowed = AM_SH                  # {"A"}
    elif clean == "P":
        allowed = PM_SH                  # {"P"}
    elif clean == "AP":
        allowed = AM_SH | PM_SH     # {"A","P"}
    elif clean == "noA":
        # allow everything EXCEPT AM shifts
        allowed = set(SHIFTS) - AM_SH
    elif clean == "noP":
        # allow everything EXCEPT PM shifts
        allowed = set(SHIFTS) - PM_SH
    elif clean == "noN":
        # allow everything EXCEPT Night shifts
        allowed = set(SHIFTS) - NIGHT_SH
    elif clean == "O":
        allowed = OFF                    # {"O","AL","☆","½●½O","兒","父","SH","PH"}
    elif clean == "☆":
        allowed = {"☆"}
    elif clean == "AL":
        allowed = {"AL"}
    elif clean == "½●½O":
        allowed = {"½●½O"}
    elif clean == "兒":
        allowed = {"兒"}
    elif clean == "父":
        allowed = {"父"}
    elif clean == "SH":
        allowed = {"SH"}
    elif clean == "PH":
        allowed = {"PH"}
    elif clean == "B":
        allowed = {"A"}
    elif clean == "N*":
        allowed = {"N*"}
    elif clean == "Z":
        allowed = {"Z"}
    elif clean == "T":
        allowed = {"T"}
    elif clean == "½t":
        allowed = {"½t"}
    elif clean == "½Z":
        allowed = {"½Z"}
    elif clean == "N3":
        allowed = {"N3"}
    elif clean == "N":
        allowed = NIGHT_SH               # {"N*","N","N3"}
    else:
        allowed = OFF                    # fallback: any OFF-type

    # exactly one of allowed…
    model.Add(sum(X[s, d, t] for t in allowed) == 1)
    # …and forbid everything else
    for t in SHIFTS:
        if t not in allowed:
            model.Add(X[s, d, t] == 0)


#### CONSTRAINTS ####



# daily coverage & ≤3 Z2
for d in range(NUM_DAYS):
    am, pm, nt = cov_targets_per_day[d]
    model.Add(sum(X[s,d,t] for s in range(NUM_STAFF) for t in AM_SH) == am)
    model.Add(sum(X[s,d,t] for s in range(NUM_STAFF) for t in PM_SH) == pm)
    model.Add(sum(X[s,d,t] for s in range(NUM_STAFF) for t in NIGHT_SH) == nt)
    # model.Add(sum(X[s,d,"Z"] for s in range(NUM_STAFF)) <= 3)

# --- Avoid PM→PM across Saturday→Sunday ---
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS-1):
        if weekdays[d]=="SAT" and weekdays[d+1]=="SUN":
            # at most one PM in that pair
            for t in PM_SH:
                model.Add(X[s,d,t] + X[s,d+1,t] <= 1)

# --- Count 3-consecutive PM runs ---
three_pm = []
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS-2):
        v = model.NewBoolVar(f"pm3_{s}_{d}")
        # v == 1 ↔ staff s has PM on d,d+1,d+2
        model.AddBoolAnd([X[s,d,t] for t in PM_SH] +
                         [X[s,d+1,t] for t in PM_SH] +
                         [X[s,d+2,t] for t in PM_SH]).OnlyEnforceIf(v)
        model.AddBoolOr([X[s,d,t].Not() for t in PM_SH] +
                        [X[s,d+1,t].Not() for t in PM_SH] +
                        [X[s,d+2,t].Not() for t in PM_SH] +
                        [v])
        three_pm.append(v)

# --- Count 4-consecutive PM runs ---
four_pm = []
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS - 3):
        v = model.NewBoolVar(f"pm4_{s}_{d}")
        model.AddBoolAnd([X[s, d+k, t] for k in range(4) for t in PM_SH]).OnlyEnforceIf(v)
        model.AddBoolOr([X[s, d+k, t].Not() for k in range(4) for t in PM_SH] + [v])
        four_pm.append(v)

# --- Count PM→AM→Night chains ---
pm_am_n = []
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS-2):
        v = model.NewBoolVar(f"chain_{s}_{d}")
        model.AddBoolAnd(
            [X[s,d,t] for t in PM_SH] +
            [X[s,d+1,t] for t in AM_SH] +
            [X[s,d+2,t] for t in NIGHT_SH]
        ).OnlyEnforceIf(v)
        model.AddBoolOr(
            [x.Not() for x in ([X[s,d,t] for t in PM_SH] +
                               [X[s,d+1,t] for t in AM_SH] +
                               [X[s,d+2,t] for t in NIGHT_SH])] + [v]
        )
        pm_am_n.append(v)
        
# quotas & one N*/day
for s,name in enumerate(staff):
    model.Add(sum(X[s,d,"N*"] for d in range(NUM_DAYS)) == quotas[name]["N*"])
    model.Add(sum(X[s,d,"N"]   for d in range(NUM_DAYS)) == quotas[name]["N"])
    model.Add(sum(X[s,d,"N3"]   for d in range(NUM_DAYS)) == quotas[name]["N3"])
#    model.Add(sum(X[s,d,"Z"]  for d in range(NUM_DAYS)) == quotas[name]["Z"])
for d in range(NUM_DAYS):
    model.Add(sum(X[s,d,"N*"] for s in range(NUM_STAFF)) == 1)
    model.Add(sum(X[s,d,"N"] for s in range(NUM_STAFF)) == 1)
   
# --- Exact SUN-Off & WE-Off quotas (only if given) ---
for s, name in enumerate(staff):
    sun_req = quotas[name]["SUN-Off"]
    we_req  = quotas[name]["WE-Off"]

    if sun_req is not None:
        # exactly sun_req OFF (any code in OFF) on Sundays
        model.Add(
            sum(
                sum(X[s, d, t] for t in OFF)
                for d, wd in enumerate(weekdays)
                if wd == "SUN"
            ) == sun_req
        )

    if we_req is not None:
        # exactly we_req OFF on any WE‐type day (weekend/holiday)
        model.Add(
            sum(
                sum(X[s, d, t] for t in OFF)
                for d in range(NUM_DAYS)
                if day_type[d] == "WE"
            ) == we_req
        )


# === Adjacency & N-spacing (FORBID Night after A↗) ===
# Uses: OFF, NIGHT_SH, fixed_raw, fixed_clean, prev_last7, prev_last7_arrow

for s in range(NUM_STAFF):
    for d in range(NUM_DAYS - 1):
        # 1) After any night, next day must be OFF
        for nt in NIGHT_SH:
            model.Add(sum(X[s, d+1, t] for t in OFF) == 1).OnlyEnforceIf(X[s, d, nt])

        # 2) If today is a fixed A↗, FORBID Night tomorrow
        is_fixed_A_arrow = (
            (s, d) in fixed_raw
            and str(fixed_clean[(s, d)]).strip().upper() == "A"
            and str(fixed_raw[(s, d)]).endswith("↗")
        )
        if is_fixed_A_arrow:
            model.Add(sum(X[s, d+1, nt] for nt in NIGHT_SH) == 0)
        else:
            # Otherwise: if tomorrow is Night, today must be plain A
            for nt in NIGHT_SH:
                model.Add(X[s, d, "A"] == 1).OnlyEnforceIf(X[s, d+1, nt])

    # === Cross-month adjacency (day -1 → day 0) ===
    prev_core  = str(prev_last7[s][6] or "").strip()
    prev_arrow = bool(prev_last7_arrow[s][6])

    # If last day of last month was a Night, day 0 must be OFF
    if prev_core in NIGHT_SH:
        model.Add(sum(X[s, 0, t] for t in OFF) == 1)

    # If last day of last month was A↗, FORBID Night on day 0
    if (prev_core == "A") and prev_arrow:
        for nt in NIGHT_SH:
            model.Add(X[s, 0, nt] == 0)
    else:
        # Otherwise: allow Night on day 0 only if yesterday was plain A
        is_prev_A_plain = (prev_core == "A" and not prev_arrow)
        if not is_prev_A_plain:
            for nt in NIGHT_SH:
                model.Add(X[s, 0, nt] == 0)

    # === N-spacing (min interval between nights), includes cross-day within month
    name = staff[s]
    indiv_spacing = int(quotas[name].get("Nspacing", 0))
    spacing = max(indiv_spacing, global_Nspacing)

    if spacing > 1:
        for d1 in range(NUM_DAYS):
            stop = min(d1 + spacing, NUM_DAYS)  # nights forbidden on d2 in (d1, stop)
            for d2 in range(d1 + 1, stop):
                model.Add(
                    sum(X[s, d1, nt] for nt in NIGHT_SH) +
                    sum(X[s, d2, nt] for nt in NIGHT_SH)
                    <= 1
                )


# === N spacing including cross-month boundary ===
for s in range(NUM_STAFF):
    name = staff[s]
    indiv_spacing = int(quotas[name].get("Nspacing", 0))   # per-doctor (col I)
    spacing = max(indiv_spacing, global_Nspacing)

    if spacing > 1:
        # 1) Intra-month: already implemented
        for d1 in range(NUM_DAYS):
            stop = min(d1 + spacing, NUM_DAYS)
            for d2 in range(d1 + 1, stop):
                model.Add(
                    sum(X[s, d1, nt] for nt in NIGHT_SH) +
                    sum(X[s, d2, nt] for nt in NIGHT_SH)
                    <= 1
                )

        # 2) Cross-month: check last 7 days of previous month
        for offset, prev_core in enumerate(prev_last7[s]):
            if str(prev_core).strip() in NIGHT_SH:
                prev_day_index = -7 + offset  # -7 .. -1

                # distance into this month must respect spacing
                for d in range(min(spacing, NUM_DAYS)):
                    # day index in this month
                    if d < spacing - abs(prev_day_index):   # ensure correct gap
                        for nt in NIGHT_SH:
                            model.Add(X[s, d, nt] == 0)
                            
# COS: follow original requests only (no auto-assignment)
for s, name in enumerate(staff):
    if ranks[s].upper().startswith("COS"):
        for d in range(NUM_DAYS):
            if (s, d) in fixed_clean:
                # Prefilled request already enforced by your fixed-override logic.
                # Do not add anything here to avoid conflicts.
                pass
            else:
                 # No request -> force an OFF-type shift (any of OFF set)
                 model.Add(sum(X[s, d, t] for t in OFF) == 1)
                 for t in set(SHIFTS) - OFF:
                    model.Add(X[s, d, t] == 0)

# CON1 & CON2, weekly Z <= 3  (prefill-aware; OFF includes ☆)
# Make sure OFF is defined earlier like: OFF = {"O","AL","☆","½●½O","兒","父" "SH","PH"}
for s, name in enumerate(staff):
    if ranks[s].upper().startswith(("CON1", "CON2")):
        for d in range(NUM_DAYS):
            # If a duty was prefilled for this (staff, day), let that override entirely.
            if (s, d) in fixed_clean:
                continue  # prefilled logic already enforced elsewhere

            if day_type[d] == "WE":
                # Weekend/holiday must be OFF (any OFF subtype allowed, incl ☆)
                model.Add(sum(X[s, d, t] for t in OFF) == 1)
                # Forbid any non-OFF on WE
                for t in set(SHIFTS) - OFF:
                    model.Add(X[s, d, t] == 0)
            else:
                # On non-WE days, allow only A, Z, or OFF; forbid everything else
                allowed = {"A", "Z"} | OFF
                for t in set(SHIFTS) - allowed:
                    model.Add(X[s, d, t] == 0)

        # Weekly Z cap (non-overlapping 7-day chunks)
        for wk in range(0, NUM_DAYS, 7):
            model.Add(sum(X[s, d, "Z"] for d in range(wk, min(wk + 7, NUM_DAYS))) <= 5)


# --- Enforce per-day min/max by cadre on A and P ---
for d in range(NUM_DAYS):
    am_cov, pm_cov, _nt_cov = cov_targets_per_day[d]

    # Clamp bounds to daily coverage so we don't create impossible constraints
    minsenA = min(int(minsenA_per_day[d] or 0), am_cov)
    maxsenA = min(int(maxsenA_per_day[d] or 99), am_cov)
    minsenP = min(int(minsenP_per_day[d] or 0), pm_cov)
    maxsenP = min(int(maxsenP_per_day[d] or 99), pm_cov)

    # Seniors on A/P
    model.Add(sum(X[s, d, "A"] for s in senior_idx) >= minsenA)
    model.Add(sum(X[s, d, "A"] for s in senior_idx) <= maxsenA)
    model.Add(sum(X[s, d, "P"] for s in senior_idx) >= minsenP)
    model.Add(sum(X[s, d, "P"] for s in senior_idx) <= maxsenP)


# --- Enforce: each day ≥1 shift‐specialist on AM not followed by night next day ---
# Define your shift‐specialists set if not already defined:
shift_specialists = {
    s for s, r in enumerate(ranks)
    if r.upper().startswith(("CON1","CON2","CON3","AC"))
}

# Build helper BoolVars for “AM today and no night tomorrow”
spec_ok = {}
for s in shift_specialists:
    for d in range(NUM_DAYS - 1):
        v = model.NewBoolVar(f"spec_ok_{s}_{d}")
        # v == 1 ↔ X[s,d,"A"] is true AND X[s,d+1,night] are all false
        model.AddBoolAnd(
            [X[s, d, "A"]] +
            [X[s, d+1, nt].Not() for nt in NIGHT_SH]
        ).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, d, "A"].Not()] +
            [X[s, d+1, nt]    for nt in NIGHT_SH] +
            [v]
        )
        spec_ok[(s, d)] = v

# Finally, require at least one such specialist each day
for d in range(NUM_DAYS - 1):
    model.Add(
        sum(v for (s, dd), v in spec_ok.items() if dd == d)
        >= 1
    )


# cap consecutive working days to ≤6
for s in range(NUM_STAFF):
    for start in range(NUM_DAYS - 6):
        # sum of non-OFF in any 7-day window ≤ 6
        model.Add(sum(
            X[s, d, t]
            for d in range(start, start + 7)
            for t in SHIFTS if t not in OFF
        ) <= 6)

# === Extend "≤6 work days in any 7-day window" across the month boundary ===
# We treat any non-OFF code in prev_last7 as a "work day".
# prev_last7[s][0] ... prev_last7[s][6]  → 7th to 1st day BEFORE day 1 of this month
# Index 6 is the last day of last month.

def _is_work(code: str) -> int:
    return 0 if (str(code or "").strip() in OFF) else 1

for s in range(NUM_STAFF):
    # For k = 1..6, consider the 7-day window that ends before we have
    # k previous days and (7-k) current days starting at day 0.
    # Example: k=2 window = [prev_day(-2), prev_day(-1), day0, day1, day2, day3, day4]
    for k in range(1, 7):
        # Count how many of the last k previous days were "work"
        # These are indices [7-k, 6] in prev_last7
        prev_work_cnt = sum(_is_work(prev_last7[s][i]) for i in range(7 - k, 7))

        # Number of current days in this cross-boundary window
        cur_len = 7 - k
        if cur_len <= 0:
            continue  # no current days to constrain

        # Upper bound for "work" among the first cur_len days of this month
        # to keep total ≤ 6 in the 7-day window
        allow_cur_work = 6 - prev_work_cnt
        if allow_cur_work < 0:
            # Already violated in prior month; the best we can do now is
            # force the current part to 0 work for this window.
            allow_cur_work = 0

        model.Add(
            sum(
                X[s, d, t]
                for d in range(cur_len)               # days 0 .. (cur_len-1)
                for t in SHIFTS if t not in OFF
            ) <= allow_cur_work
        )

        


# 1) Build the HT1+HT2 index set
ht_group = {
    s for s, r in enumerate(ranks)
    if r.upper().startswith(("HT1", "HT2"))
}

# 2) For each day classified as "WE", require at least one HT on AM
for d in range(NUM_DAYS):
    if day_type[d] == "WE":
        model.Add(
            sum(X[s, d, "A"] for s in ht_group)
            >= 1
        )


# 2.x Hours & balance constraints
threshold = int(ws_roster["O2"].value or 0)

# define shift durations
shift_hours = {t: 9 for t in SHIFTS}
shift_hours["O"] = 0

hour_assigned = {}
final_hr      = {}
non_cos       = [s for s,name in enumerate(staff)
                 if not ranks[s].upper().startswith("COS")]

# --- NEW: read per-staff target ranges (min/max) from CF/CG ---
# CF = 84 (min), CG = 85 (max)
per_staff_bounds = {}  # s -> (min_or_None, max_or_None)

for s, name in enumerate(staff):
    row = ROW_START + s
    raw_min = ws_roster.cell(row=row, column=16).value  # CF
    raw_max = ws_roster.cell(row=row, column=17).value  # CG

    # Clean cast to int if present, else None
    lb = None if raw_min in (None, "") else int(raw_min)
    ub = None if raw_max in (None, "") else int(raw_max)

    # If someone accidentally types lb>ub, swap to be safe
    if (lb is not None) and (ub is not None) and (lb > ub):
        lb, ub = ub, lb

    per_staff_bounds[s] = (lb, ub)

for s,name in enumerate(staff):
    # total hours assigned (exclude prefilled requests)
    h = model.NewIntVar(0, NUM_DAYS*9, f"hrs_{s}")
    hour_assigned[s] = h
    model.Add(
        h == sum(
            X[s,d,t] * shift_hours[t]
            for d in range(NUM_DAYS)
            for t in SHIFTS
            if (s,d) not in fixed_clean  # skip fixed requests
        )
    )

    # final hour balance
    fh = model.NewIntVar(-100000, 100000, f"fh_{s}")
    final_hr[s] = fh
    model.Add(fh == int(quotas[name]["init_hr"]) + h)

    # --- NEW: enforce personal min/max if provided ---
    lb, ub = per_staff_bounds[s]
    if lb is not None:
        model.Add(fh >= lb)
    if ub is not None:
        model.Add(fh <= ub)

# --- Pairwise hard cap (only among staff WITHOUT a personal range) ---
eligible_for_pairwise = [
    s for s in non_cos
    if per_staff_bounds[s][0] is None and per_staff_bounds[s][1] is None
]

for i in range(len(eligible_for_pairwise)):
    for j in range(i+1, len(eligible_for_pairwise)):
        si, sj = eligible_for_pairwise[i], eligible_for_pairwise[j]
        # |final_hr[si] - final_hr[sj]| <= threshold
        diff = model.NewIntVar(-threshold, threshold, f"diff_{si}_{sj}")
        model.Add(diff == final_hr[si] - final_hr[sj])
        absdiff = model.NewIntVar(0, threshold, f"abs_{si}_{sj}")
        model.AddAbsEquality(absdiff, diff)



# === P/A‐ratio constraints (per staff) ===
for s, name in enumerate(staff):
    # count PM and AM
    total_pm = sum(X[(s, d, t)] for d in range(NUM_DAYS) for t in PM_SH)
    total_am = sum(X[(s, d, t)] for d in range(NUM_DAYS) for t in AM_SH)

    # retrieve from quotas
    max_pa = int(quotas[name]["max_pa"] * 100)
    min_pa = int(quotas[name]["min_pa"] * 100)

    # enforce: min_pa ≤ total_pm/total_am ≤ max_pa
    model.Add(total_pm * 100 <= total_am * max_pa)
    model.Add(total_pm * 100 >= total_am * min_pa)

# === Global P/A ratio constraint ===
total_pm_all = sum(X[(s, d, t)] for s in range(NUM_STAFF) for d in range(NUM_DAYS) for t in PM_SH)
total_am_all = sum(X[(s, d, t)] for s in range(NUM_STAFF) for d in range(NUM_DAYS) for t in AM_SH)

scaling_factor = 100
globalmax_pa = int(globalmax_pa * scaling_factor)  # Make sure it's defined or read from Excel

# Enforce: total_pm_all / total_am_all ≤ globalmax_pa
model.Add(total_pm_all * scaling_factor <= total_am_all * globalmax_pa)



# === 2.x Build penalty variables for soft objectives ===
# --- PM→AM→Night detectors (PAN) ---
# --- Unified PAN detector: count N-days that have PA immediately before (handles month boundary) ---
pm_am_night = []

def _norm(code: object) -> str:
    # normalize prior-month codes like "P*" -> "P", "A*" -> "A"
    return str(code or "").strip().rstrip("*")

for s in range(NUM_STAFF):
    for d in range(NUM_DAYS):
        # isNight(d)
        isN = model.NewBoolVar(f"isN_{s}_{d}")
        model.Add(sum(X[s, d, t] for t in NIGHT_SH) == 1).OnlyEnforceIf(isN)
        model.Add(sum(X[s, d, t] for t in NIGHT_SH) == 0).OnlyEnforceIf(isN.Not())

        # prev is AM: day d-1 if in-month, else last day of prev month
        prevAM = model.NewBoolVar(f"prevAM_{s}_{d}")
        if d >= 1:
            model.Add(sum(X[s, d-1, t] for t in AM_SH) == 1).OnlyEnforceIf(prevAM)
            model.Add(sum(X[s, d-1, t] for t in AM_SH) == 0).OnlyEnforceIf(prevAM.Not())
        else:
            # d == 0 → use prev_last7[s][6]
            model.Add(prevAM == (1 if _norm(prev_last7[s][6]) == "A" else 0))

        # prevprev is PM: day d-2 if in-month, else pull from prev month as needed
        prevPM = model.NewBoolVar(f"prevPM_{s}_{d}")
        if d >= 2:
            model.Add(sum(X[s, d-2, t] for t in PM_SH) == 1).OnlyEnforceIf(prevPM)
            model.Add(sum(X[s, d-2, t] for t in PM_SH) == 0).OnlyEnforceIf(prevPM.Not())
        elif d == 1:
            # look at prev_last7[s][6]
            model.Add(prevPM == (1 if _norm(prev_last7[s][6]) == "P" else 0))
        else:  # d == 0 → look at prev_last7[s][5]
            model.Add(prevPM == (1 if _norm(prev_last7[s][5]) == "P" else 0))

        # v == 1  <=>  (prevprev=PM) ∧ (prev=AM) ∧ (today=Night)
        v = model.NewBoolVar(f"pan_hit_{s}_{d}")
        model.AddBoolAnd([prevPM, prevAM, isN]).OnlyEnforceIf(v)
        model.AddBoolOr([prevPM.Not(), prevAM.Not(), isN.Not(), v])

        pm_am_night.append(v)

pm_am = []
# PM→AM transitions (PA)
for s in range(NUM_STAFF):
    for d in range(1, NUM_DAYS):
        v = model.NewBoolVar(f"pm_am_{s}_{d}")
        model.AddBoolAnd(
            [X[s, d-1, t] for t in PM_SH] +
            [X[s, d,   t] for t in AM_SH]
        ).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, d-1, t].Not() for t in PM_SH] +
            [X[s, d,   t].Not() for t in AM_SH] +
            [v]
        )
        pm_am.append(v)

three_pm = []
# 3-consecutive PM runs
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS-2):
        v = model.NewBoolVar(f"three_pm_{s}_{d}")
        model.Add(
            sum(X[s, d+k, t] for k in range(3) for t in PM_SH) == 3
        ).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, d+k, t].Not() for k in range(3) for t in PM_SH] + [v]
        )
        three_pm.append(v)

# --- track 4-consecutive PM runs (PPPP) ---
four_pm = []
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS - 3):
        v = model.NewBoolVar(f"four_pm_{s}_{d}")
        # v == 1  <=>  exactly 4 PMs on d..d+3
        model.Add(
            sum(X[s, d+k, t] for k in range(4) for t in PM_SH) == 4
        ).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, d+k, t].Not() for k in range(4) for t in PM_SH] + [v]
        )
        four_pm.append(v)

# === Boundary-aware caps for PA / PAN / PPP (count if the last day is in THIS month) ===
# Assumes:
#   - pm_am:      intra-month PA booleans with names containing f"_{s}_"
#   - pm_am_night:intra-month PAN booleans with names containing f"_{s}_"
#   - three_pm:   intra-month PPP booleans with names containing f"_{s}_"
#   - prev_last7[s][i]: clean codes for prior month (i=6 is last day, i=5 is -2)
#   - AM_SH, PM_SH, NIGHT_SH sets; X[s,d,t] decision vars

def _norm(code: object) -> str:
    # normalize prior-month codes such as "A↗", "P*" → "A","P"
    return str(code or "").strip().rstrip("↗").rstrip("*")

def _in_prev(code_set, s, idx):
    return _norm(prev_last7[s][idx]) in code_set

def _vars_for_staff(var_list, s):
    tag = f"_{s}_"
    return [v for v in var_list if tag in v.Name()]

# Boundary detector lists (one BoolVar per staff for each boundary case)
PA_b0   = []  # PA with last day = day 0 (prev[-1]=P, day0=A)
PAN_b0  = []  # PAN with last day = day 0 (prev[-2]=P, prev[-1]=A, day0=Night)
PAN_b1  = []  # PAN with last day = day 1 (prev[-1]=P, day0=A, day1=Night)
PPP_b0  = []  # PPP with last day = day 0 (prev[-2]=P, prev[-1]=P, day0=P)
PPP_b1  = []  # PPP with last day = day 1 (prev[-1]=P, day0=P, day1=P)

for s in range(NUM_STAFF):
    # ---- PA boundary (last day d=0) ----
    if _in_prev(PM_SH, s, 6):
        v = model.NewBoolVar(f"PA_boundary_{s}_d0")
        model.AddBoolAnd([X[s, 0, t] for t in AM_SH]).OnlyEnforceIf(v)
        model.AddBoolOr([X[s, 0, t].Not() for t in AM_SH] + [v])
        PA_b0.append(v)
    else:
        # make a 0-constant Bool for uniform summation
        v = model.NewBoolVar(f"PA_boundary_{s}_d0_zero")
        model.Add(v == 0)
        PA_b0.append(v)

    # ---- PAN boundary (last day d=0) ----
    if _in_prev(PM_SH, s, 5) and _in_prev(AM_SH, s, 6):
        v = model.NewBoolVar(f"PAN_boundary_{s}_d0")
        model.AddBoolAnd([X[s, 0, t] for t in NIGHT_SH]).OnlyEnforceIf(v)
        model.AddBoolOr([X[s, 0, t].Not() for t in NIGHT_SH] + [v])
        PAN_b0.append(v)
    else:
        v = model.NewBoolVar(f"PAN_boundary_{s}_d0_zero")
        model.Add(v == 0)
        PAN_b0.append(v)

    # ---- PAN boundary (last day d=1) ----
    if _in_prev(PM_SH, s, 6) and NUM_DAYS >= 2:
        v = model.NewBoolVar(f"PAN_boundary_{s}_d1")
        # day0=A and day1=Night
        model.AddBoolAnd(
            [X[s, 0, t] for t in AM_SH] +
            [X[s, 1, t] for t in NIGHT_SH]
        ).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, 0, t].Not() for t in AM_SH] +
            [X[s, 1, t].Not() for t in NIGHT_SH] + [v]
        )
        PAN_b1.append(v)
    else:
        v = model.NewBoolVar(f"PAN_boundary_{s}_d1_zero")
        model.Add(v == 0)
        PAN_b1.append(v)

    # ---- PPP boundary (last day d=0) ----
    if _in_prev(PM_SH, s, 5) and _in_prev(PM_SH, s, 6):
        v = model.NewBoolVar(f"PPP_boundary_{s}_d0")
        model.Add(sum(X[s, 0, t] for t in PM_SH) == 1).OnlyEnforceIf(v)
        model.AddBoolOr([X[s, 0, t].Not() for t in PM_SH] + [v])
        PPP_b0.append(v)
    else:
        v = model.NewBoolVar(f"PPP_boundary_{s}_d0_zero")
        model.Add(v == 0)
        PPP_b0.append(v)

    # ---- PPP boundary (last day d=1) ----
    if _in_prev(PM_SH, s, 6) and NUM_DAYS >= 2:
        v = model.NewBoolVar(f"PPP_boundary_{s}_d1")
        model.Add(
            sum(X[s, 0, t] for t in PM_SH) +
            sum(X[s, 1, t] for t in PM_SH) == 2
        ).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, 0, t].Not() for t in PM_SH] +
            [X[s, 1, t].Not() for t in PM_SH] + [v]
        )
        PPP_b1.append(v)
    else:
        v = model.NewBoolVar(f"PPP_boundary_{s}_d1_zero")
        model.Add(v == 0)
        PPP_b1.append(v)

# Build per-doctor IntVars = intra-month detectors + boundary detectors
PA_cnt_var  = {}
PAN_cnt_var = {}
PPP_cnt_var = {}

def _as_int_or(v, default=99):
    try: return int(v)
    except (TypeError, ValueError): return default

for s in range(NUM_STAFF):
    pa_vars_in   = _vars_for_staff(pm_am,       s)  # includes any in-month PA vars
    pan_vars_in  = _vars_for_staff(pm_am_night, s)  # includes in-month PAN vars
    ppp_vars_in  = _vars_for_staff(three_pm,    s)  # includes in-month PPP vars

    # Add matching boundary var for this s (same index in the lists)
    vPA  = model.NewIntVar(0, len(pa_vars_in)+1,             f"PA_sum_{s}")
    vPAN = model.NewIntVar(0, len(pan_vars_in)+2,            f"PAN_sum_{s}")
    vPPP = model.NewIntVar(0, len(ppp_vars_in)+2,            f"PPP_sum_{s}")

    # Each boundary list is in staff order; pick the s-th entry
    model.Add(vPA  == sum(pa_vars_in)  + PA_b0[s])
    model.Add(vPAN == sum(pan_vars_in) + PAN_b0[s] + PAN_b1[s])
    model.Add(vPPP == sum(ppp_vars_in) + PPP_b0[s] + PPP_b1[s])

    PA_cnt_var[s]  = vPA
    PAN_cnt_var[s] = vPAN
    PPP_cnt_var[s] = vPPP

    # Caps from R,S,T (blank -> 99)
    row = ROW_START + s
    cap_PA  = _as_int_or(ws_roster.cell(row=row, column=18).value, 99)  # R
    cap_PAN = _as_int_or(ws_roster.cell(row=row, column=19).value, 99)  # S
    cap_PPP = _as_int_or(ws_roster.cell(row=row, column=20).value, 99)  # T

    model.Add(vPA  <= cap_PA)
    model.Add(vPAN <= cap_PAN)
    model.Add(vPPP <= cap_PPP)


# === Read single toggle from D3: apply ALL penalties if "Y" ===
do_penalty = str(ws_roster.cell(row=3, column=4).value or "").strip().upper() == "Y"

# === Cross‑month penalty detectors (extend penalties across the boundary) ===
# Uses prev_last7[s][i] (clean code) where i=6 is last day of last month.
# Reuses AM_SH, PM_SH, NIGHT_SH, and the penalty arrays you already built.

def _in_set(code, sset):
    return str(code or "").strip() in sset

for s in range(NUM_STAFF):
    # ---------- PA (PM → AM) across boundary: [-1, 0] ----------
    if _in_set(prev_last7[s][6], PM_SH):
        v = model.NewBoolVar(f"pm_am_boundary_{s}_0")
        # v == 1  <=>  day0 is AM  (prev day is fixed PM, so it's a constant condition)
        model.AddBoolAnd([X[s, 0, t] for t in AM_SH]).OnlyEnforceIf(v)
        model.AddBoolOr([X[s, 0, t].Not() for t in AM_SH] + [v])
        pm_am.append(v)

    # ---------- PAN (PM → AM → Night) across boundary ----------
    # Case A: [-2, -1, 0]  (prev[-2]=PM, prev[-1]=AM, day0=Night)
    if _in_set(prev_last7[s][5], PM_SH) and _in_set(prev_last7[s][6], AM_SH):
        v = model.NewBoolVar(f"pan_boundary_A_{s}_0")
        model.AddBoolAnd([X[s, 0, t] for t in NIGHT_SH]).OnlyEnforceIf(v)
        model.AddBoolOr([X[s, 0, t].Not() for t in NIGHT_SH] + [v])
        pm_am_night.append(v)

    # Case B: [-1, 0, 1]  (prev[-1]=PM, day0=AM, day1=Night)
    if _in_set(prev_last7[s][6], PM_SH) and NUM_DAYS >= 2:
        v = model.NewBoolVar(f"pan_boundary_B_{s}_0_1")
        model.AddBoolAnd(
            [X[s, 0, t] for t in AM_SH] +
            [X[s, 1, t] for t in NIGHT_SH]
        ).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, 0, t].Not() for t in AM_SH] +
            [X[s, 1, t].Not() for t in NIGHT_SH] + [v]
        )
        pm_am_night.append(v)

    # ---------- 3×PM across boundary ----------
    # Patterns: [-2,-1,0], [-1,0,1]
    # a) [-2,-1,0] = prev[-2]=PM, prev[-1]=PM, day0=PM
    if _in_set(prev_last7[s][5], PM_SH) and _in_set(prev_last7[s][6], PM_SH):
        v = model.NewBoolVar(f"three_pm_boundary_{s}_m2_m1_0")
        model.Add(sum(X[s, 0, t] for t in PM_SH) == 1).OnlyEnforceIf(v)
        model.AddBoolOr([X[s, 0, t].Not() for t in PM_SH] + [v])
        three_pm.append(v)

    # b) [-1,0,1] = prev[-1]=PM, day0=PM, day1=PM
    if _in_set(prev_last7[s][6], PM_SH) and NUM_DAYS >= 2:
        v = model.NewBoolVar(f"three_pm_boundary_{s}_m1_0_1")
        model.Add(sum(X[s, 0, t] for t in PM_SH) +
                  sum(X[s, 1, t] for t in PM_SH) == 2).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, 0, t].Not() for t in PM_SH] +
            [X[s, 1, t].Not() for t in PM_SH] + [v]
        )
        three_pm.append(v)

    # ---------- 4×PM across boundary (only if you track it) ----------
    # Patterns: [-3,-2,-1,0], [-2,-1,0,1], [-1,0,1,2]
    # a) [-3,-2,-1,0]
    if (_in_set(prev_last7[s][4], PM_SH) and
        _in_set(prev_last7[s][5], PM_SH) and
        _in_set(prev_last7[s][6], PM_SH)):
        v = model.NewBoolVar(f"four_pm_boundary_{s}_m3_m2_m1_0")
        model.Add(sum(X[s, 0, t] for t in PM_SH) == 1).OnlyEnforceIf(v)
        model.AddBoolOr([X[s, 0, t].Not() for t in PM_SH] + [v])
        four_pm.append(v)

    # b) [-2,-1,0,1]
    if (_in_set(prev_last7[s][5], PM_SH) and
        _in_set(prev_last7[s][6], PM_SH) and NUM_DAYS >= 2):
        v = model.NewBoolVar(f"four_pm_boundary_{s}_m2_m1_0_1")
        model.Add(sum(X[s, 0, t] for t in PM_SH) +
                  sum(X[s, 1, t] for t in PM_SH) == 2).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, 0, t].Not() for t in PM_SH] +
            [X[s, 1, t].Not() for t in PM_SH] + [v]
        )
        four_pm.append(v)

    # c) [-1,0,1,2]
    if _in_set(prev_last7[s][6], PM_SH) and NUM_DAYS >= 3:
        v = model.NewBoolVar(f"four_pm_boundary_{s}_m1_0_1_2")
        model.Add(sum(X[s, 0, t] for t in PM_SH) +
                  sum(X[s, 1, t] for t in PM_SH) +
                  sum(X[s, 2, t] for t in PM_SH) == 3).OnlyEnforceIf(v)
        model.AddBoolOr(
            [X[s, 0, t].Not() for t in PM_SH] +
            [X[s, 1, t].Not() for t in PM_SH] +
            [X[s, 2, t].Not() for t in PM_SH] + [v]
        )
        four_pm.append(v)

# === Penalty Weights / Toggles ===
# Use integer for soft penalty, "X" (string) to prohibit pattern
W_PAN = 10000   # PM→AM→Night chains
W_PA  = 1000   # PM→AM transitions
W_3PM = 1000   # 3 consecutive PMs
W_4PM = "X"    # 4 consecutive PMs (here: prohibit)

# === Objective / Hard Constraints ===
if do_penalty:
    obj_terms = []

    # PAN (PM→AM→Night)
    if W_PAN == "X":
        model.Add(sum(pm_am_night) == 0)   # forbid
    else:
        obj_terms.append(W_PAN * sum(pm_am_night))

    # PA (PM→AM)
    if W_PA == "X":
        model.Add(sum(pm_am) == 0)         # forbid
    else:
        obj_terms.append(W_PA * sum(pm_am))

    # 3×PM
    if W_3PM == "X":
        model.Add(sum(three_pm) == 0)      # forbid
    else:
        obj_terms.append(W_3PM * sum(three_pm))

    # 4×PM
    if W_4PM == "X":
        model.Add(sum(four_pm) == 0)       # forbid
    else:
        obj_terms.append(W_4PM * sum(four_pm))

    model.Minimize(sum(obj_terms))
else:
    model.Minimize(0)

##
# === Per-day PA caps (counted on the A day), reading caps from row 59 ===
# If a cell in row 59 is blank/non-numeric, treat as NUM_STAFF (no effective cap).

def _norm(code: object) -> str:
    # normalize prior-month codes like "P*", "A↗" → "P","A"
    return str(code or "").strip().rstrip("*").rstrip("↗")

# Reify helper (var ↔ (assigned shift for (s,d) is in SET))
def _reify_in_set(var, s, d, SET):
    model.Add(sum(X[s, d, t] for t in SET) == 1).OnlyEnforceIf(var)
    model.Add(sum(X[s, d, t] for t in SET) == 0).OnlyEnforceIf(var.Not())

# Build PA-on-day booleans: pa_on_day[(s,d)] == 1 iff (d-1 is PM, d is AM).
pa_on_day = {}

for s in range(NUM_STAFF):
    for d in range(NUM_DAYS):
        v = model.NewBoolVar(f"PA_on_day_{s}_{d}")  # counts on day d (the AM day)

        if d >= 1:
            isAM   = model.NewBoolVar(f"isAM_{s}_{d}")
            isPrev = model.NewBoolVar(f"isPM_{s}_{d-1}")
            _reify_in_set(isAM,   s, d,   AM_SH)
            _reify_in_set(isPrev, s, d-1, PM_SH)

            # v ↔ (isPrev ∧ isAM)
            model.AddBoolAnd([isPrev, isAM]).OnlyEnforceIf(v)
            model.AddBoolOr([isPrev.Not(), isAM.Not(), v])
        else:
            # d == 0: look at last day of previous month for PM
            if _norm(prev_last7[s][6]) == "P":
                isAM0 = model.NewBoolVar(f"isAM_{s}_0")
                _reify_in_set(isAM0, s, 0, AM_SH)
                # v == isAM0
                model.Add(v == isAM0)
            else:
                model.Add(v == 0)

        pa_on_day[(s, d)] = v

# Daily P-A, Read caps from row 68 over the day columns and enforce: sum_s pa_on_day[s,d] ≤ cap_d
for d in range(NUM_DAYS):
    cap_cell = ws_roster.cell(row=68, column=START_COL + d).value
    try:
        cap_d = int(cap_cell)
    except (TypeError, ValueError):
        cap_d = NUM_STAFF  # treat empty/invalid as no effective cap

    model.Add(sum(pa_on_day[(s, d)] for s in range(NUM_STAFF)) <= cap_d)

##    
# --- OFF / Sunday-P constraints (with min & max bounds) ---
for s, name in enumerate(staff):
    # --- Sundays: OFF count (any OFF-type) ---
    sun_off = sum(
        sum(X[s, d, t] for t in OFF)
        for d, wd in enumerate(weekdays)
        if wd == "SUN"
    )
    model.Add(sun_off >= min_sun_off)
    model.Add(sun_off <= max_sun_off)

    # --- Weekend/holiday (WE): OFF count (any OFF-type) ---
    we_off = sum(
        sum(X[s, d, t] for t in OFF)
        for d in range(NUM_DAYS)
        if day_type[d] == "WE"
    )
    model.Add(we_off >= min_we_off)
    model.Add(we_off <= max_we_off)

    # --- Sunday PM duties ---
    sun_pm = sum(
        X[s, d, "P"]
        for d, wd in enumerate(weekdays)
        if wd == "SUN"
    )

    sun_p_quota = quotas[name]["SUN P"]  # exact quota if provided
    if sun_p_quota is not None:
        # Exact requirement for this doctor
        model.Add(sun_pm == sun_p_quota)
        # (Optional) keep the global upper bound too; equality will satisfy it if consistent
        model.Add(sun_pm <= max_sun_pm)
    else:
        # Otherwise, use global min/max bounds
        model.Add(sun_pm >= min_sun_pm)
        model.Add(sun_pm <= max_sun_pm)



# --- Weekly PM cap: ≤4 per staff per 7-day block starting from first Sunday ---

# 1) Find the index of the first Sunday
first_sun = next((i for i, wd in enumerate(weekdays) if wd == "SUN"), 0)

# 2) For each staff, slide a 7-day window starting at first_sun, then every 7 days
for s in range(NUM_STAFF):
    for wk_start in range(first_sun, NUM_DAYS, 7):
        wk_end = min(wk_start + 7, NUM_DAYS)
        model.Add(
            sum(X[s, d, "P"] for d in range(wk_start, wk_end))
            <= 4
        )
            
# solve

solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = 300
solver.parameters.random_seed = random.randrange(1,10000)
res = solver.Solve(model)
if res not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    print("❌ No feasible solution.")
    exit(1)

    
# === PART3. Diagnostics & Output ===

    
# write back
grey = PatternFill("solid", fgColor="C0C0C0")
red  = Font(color="FF0000")
for s in range(NUM_STAFF):
    for d in range(NUM_DAYS):
        cell = ws_roster.cell(row=ROW_START+s, column=START_COL+d)
        if (s,d) in fixed_raw:
            orig  = fixed_raw[(s,d)]
            arrow = orig.endswith("↗")
            for t in SHIFTS:
                if solver.Value(X[s,d,t]):
                    cell.value = t + ("↗" if arrow else "")
                    break
            cell.fill = grey
        else:
            for t in SHIFTS:
                if solver.Value(X[s,d,t]):
                    cell.value = t
                    break
        if day_type[d]=="WE":
            cell.font = red
# === Combined writeback anchored at "Output" ===

def _core(v):
    return str(v or "").rstrip("↗").strip()

def _find_output_col(ws, max_rows=10):
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(r, c).value or "").strip().lower()
            if val == "output":
                return c
    raise RuntimeError('Could not find a cell with text "Output" in the first rows.')

# Locate the anchor column for the metrics
OUTPUT_COL = _find_output_col(ws_roster)

for s, name in enumerate(staff):
    row = ROW_START + s

    # --- compute star hours from the sheet (prefilled/converted ☆) ---
    star_count = sum(
        1
        for d in range(NUM_DAYS)
        if _core(ws_roster.cell(row=row, column=START_COL + d).value) == "☆"
    )
    star_hours = 0 * star_count

    # --- duty hours & final hr (solver vars exclude prefilled; add star-hours) ---
    duty_hours  = solver.Value(hour_assigned[s]) + star_hours
    final_hours = solver.Value(final_hr[s]) + star_hours

    # --- P/A ratio from what’s actually written in the grid (post-solve values) ---
    pm_count = 0
    am_count = 0
    for d in range(NUM_DAYS):
        val = str(ws_roster.cell(row=row, column=START_COL + d).value or "")
        shift = val.rstrip("↗")
        if shift in PM_SH:
            pm_count += 1
        if shift in AM_SH:
            am_count += 1
    pa_ratio = (pm_count / am_count) if am_count else 0.0

    # --- Sun Off & WE Off counts (any OFF-type: OFF set) ---
    sun_off_count = sum(
        1
        for d, wd in enumerate(weekdays)
        if wd == "SUN" and _core(ws_roster.cell(row=row, column=START_COL + d).value) in OFF
    )
    we_off_count = sum(
        1
        for d in range(NUM_DAYS)
        if day_type[d] == "WE" and _core(ws_roster.cell(row=row, column=START_COL + d).value) in OFF
    )

    # --- Sun-P count ---
    sun_p_count = sum(
        1
        for d in range(NUM_DAYS)
        if weekdays[d] == "SUN"
        and str(ws_roster.cell(row=row, column=START_COL + d).value or "").rstrip("↗") == "P"
    )

    # --- Penalty counts from detector variables ---
    # PA (PM→AM), PAN (PM→AM→Night), PPP (3×PM)
    pa_cnt  = sum(solver.Value(v) for v in pm_am        if f"_{s}_" in v.Name())
    pan_cnt = sum(solver.Value(v) for v in pm_am_night  if f"_{s}_" in v.Name())
    ppp_cnt = sum(solver.Value(v) for v in three_pm     if f"_{s}_" in v.Name())

    # --- Write in required sequence starting from OUTPUT_COL ---
    col = OUTPUT_COL
    ws_roster.cell(row=row, column=col + 0).value = int(sun_off_count)       # Sun Off
    ws_roster.cell(row=row, column=col + 1).value = int(we_off_count)        # WE Off
    ws_roster.cell(row=row, column=col + 2).value = int(sun_p_count)         # Sun P
    ws_roster.cell(row=row, column=col + 3).value = round(pa_ratio, 2)       # P/A ratio
    ws_roster.cell(row=row, column=col + 4).value = int(duty_hours)          # duty hr
    ws_roster.cell(row=row, column=col + 5).value = int(final_hours)         # final hr
    ws_roster.cell(row=row, column=col + 6).value = int(pa_cnt)              # PA count
    ws_roster.cell(row=row, column=col + 7).value = int(pan_cnt)             # PAN count
    ws_roster.cell(row=row, column=col + 8).value = int(ppp_cnt)             # PPP (3×PM) count
# now save
wb_roster.save(OUTPUT_FILE)
print(f"✅ Written {OUTPUT_FILE}")



















####PART4####

# === Post‐hoc conversion block (fixed arrow handling) ===
from openpyxl import load_workbook
import random

wb2 = load_workbook("Roster_Output1.xlsx")
ws2 = wb2["Sheet1"]

# only run if D4 == "Y"
# only run if D4 == 2 or 3
try:
    toggle_val = int(ws2.cell(row=4, column=4).value or 0)
except ValueError:
    toggle_val = 0

if toggle_val < 2:
    exit(0)   # clean exit

if toggle_val >1:

    # --- helpers ---
    def split_arrow(cell):
        """Return (core, arrow) where arrow is '↗' if present."""
        s = str(cell or "")
        if s.endswith("↗"):
            return s[:-1], "↗"
        return s, ""
    
    # find START column
    c = 1
    while str(ws2.cell(1, c).value or "").strip().upper() != "START":
        c += 1
    START = c + 1

    # classify day‐cols
    day_cols, weekend, weekday = [], set(), []
    col = START
    while ws2.cell(3, col).value is not None:
        day_cols.append(col)
        wd  = str(ws2.cell(4, col).value or "").strip().upper()
        tag = str(ws2.cell(1, col).value or "").strip().upper()
        if wd in ("SAT","SUN") or tag.startswith(("PH","SH")):
            weekend.add(col)
        else:
            weekday.append(col)
        col += 1

    # collect staff rows & ranks
    ROW_START = 6
    staff_rows, ranks = [], {}
    r = ROW_START
    while True:
        nm = ws2.cell(r, 4).value
        if not nm: break
        staff_rows.append(r)
        ranks[r] = str(ws2.cell(r,3).value or "").strip().upper()
        r += 1

    # 1) N3 → N.
    for r in staff_rows:
        for c in day_cols:
            core, arr = split_arrow(ws2.cell(r,c).value)
            if core == "N3":
                ws2.cell(r,c).value = "N." + arr

    # 2) Z by CON* → Z2
    for r in staff_rows:
        if ranks[r].startswith(("COS","CON")):
            for c in day_cols:
                core, arr = split_arrow(ws2.cell(r,c).value)
                if core == "Z":
                    ws2.cell(r,c).value = "Z2" + arr
import random

# === 7) Special duties assignment ===

# 7.1 Find the "Special duties" column (row 4)
SD = None
for c in range(1, ws2.max_column + 1):
    if str(ws2.cell(4, c).value or "").strip().upper() == "SPECIAL DUTIES":
        SD = c
        break
if SD is None:
    raise RuntimeError("❌ Could not find 'Special duties' column in row 4.")

# 7.2 Read staff tiers from the SD column
tiers = {}
for r in staff_rows:
    try:
        tiers[r] = int(ws2.cell(r, SD).value or 0)
    except ValueError:
        tiers[r] = 0   # default = not eligible

# 7.3 Helper: pick a staff row by tier priority
def pick_staff_by_tier(candidates):
    # Group by tier, lowest tier first
    tier_groups = {}
    for r in candidates:
        t = tiers.get(r, 0)
        if t == 0:
            continue
        tier_groups.setdefault(t, []).append(r)
    if not tier_groups:
        return None
    for t in sorted(tier_groups.keys()):
        return random.choice(tier_groups[t])  # pick randomly within tier


NIGHT_SET = {"N*", "N", "N3", "N."}
PM_SET    = {"P", "P*", "E2", "E2*","S2","S4"}  # cover before/after bulk conversion
try:
    OFF_SET
except NameError:
    OFF_SET = {"O", "AL", "☆", "½●½O", "兒", "父", "SH", "PH"}

# === Extra module: handle P-A-N and remaining P-A transitions ===


for day_idx, c in enumerate(day_cols[:-2]):  # up to third-last day
    next_c = day_cols[day_idx + 1]
    next2_c = day_cols[day_idx + 2]

    # --- Stage 1: P-A-N transitions ---
    for r in staff_rows:
        core_today, arr_today = split_arrow(ws2.cell(r, c).value or "")
        core_next, _          = split_arrow(ws2.cell(r, next_c).value or "")
        core_next2, _         = split_arrow(ws2.cell(r, next2_c).value or "")

        if core_today == "P" and core_next == "A" and core_next2 in NIGHT_SET:
            ws2.cell(r, c).value = "S2" + arr_today  # convert P -> S2

# --- Stage 2: remaining P-A transitions ---
for day_idx, c in enumerate(day_cols[:-1]):  # up to second-last day
    next_c = day_cols[day_idx + 1]
    transition_count = 0

    for r in staff_rows:
        core_today, arr_today = split_arrow(ws2.cell(r, c).value or "")
        core_next, arr_next   = split_arrow(ws2.cell(r, next_c).value or "")

        if core_today == "P" and core_next == "A":
            transition_count += 1
            if transition_count % 2 == 1:  # odd -> convert P
                ws2.cell(r, c).value = "S2" + arr_today
            else:  # even -> convert A
                ws2.cell(r, next_c).value = "D2" + arr_next




# 7.4 IC duty For each day: assign one A→A2*, one P→E2*
for c in day_cols:
    # --- A duty ---
    cand_A = []
    for r in staff_rows:
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        # must be plain A, and tomorrow not N-type (to avoid breaking A-N-O)
        if core == "A":
            if c < day_cols[-1]:
                nxt, _ = split_arrow(ws2.cell(r, c+1).value or "")
                if nxt in ("N*", "N", "N3"):
                    continue
            cand_A.append(r)
    sel_A = pick_staff_by_tier(cand_A)
    if sel_A:
        core, arr = split_arrow(ws2.cell(sel_A, c).value or "")
        ws2.cell(sel_A, c).value = "A2*" + arr

    # --- P duty ---
    cand_P = []
    for r in staff_rows:
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if core == "P":
            cand_P.append(r)
    sel_P = pick_staff_by_tier(cand_P)
    if sel_P:
        core, arr = split_arrow(ws2.cell(sel_P, c).value or "")
        ws2.cell(sel_P, c).value = "E2*" + arr

        
# === 7b) Special duties A2^ (from SD+3 column, weekdays only) ===

SD_WD_EMW = SD + 3  # column for A2^ tiers

# 7b.1 Read staff tiers for A2^
tiers_A2hat = {}
for r in staff_rows:
    try:
        tiers_A2hat[r] = int(ws2.cell(r, SD_WD_EMW).value or 0)
    except ValueError:
        tiers_A2hat[r] = 0

# 7b.2 Helper: pick staff by tier (for A2^)
def pick_staff_by_tier_A2hat(candidates):
    tier_groups = {}
    for r in candidates:
        t = tiers_A2hat.get(r, 0)
        if t == 0:
            continue
        tier_groups.setdefault(t, []).append(r)
    if not tier_groups:
        return None
    for t in sorted(tier_groups.keys()):
        return random.choice(tier_groups[t])

# 7b.3 For each weekday: assign one A → A2^
for c in weekday:   # weekday list already built earlier
    cand_A = []
    for r in staff_rows:
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        # must be plain A, and not followed by a night shift
        if core == "A":
            if c < day_cols[-1]:
                nxt, _ = split_arrow(ws2.cell(r, c+1).value or "")
                if nxt in ("N*", "N", "N3"):
                    continue
            cand_A.append(r)
    sel_A = pick_staff_by_tier_A2hat(cand_A)
    if sel_A:
        core, arr = split_arrow(ws2.cell(sel_A, c).value or "")
        ws2.cell(sel_A, c).value = "A2^" + arr


# === Special duty B (from SD+1), strict per-tier; forbid PA; prefer A→N→… ===
SD_B = SD + 1  # column for B tiers

# Read tiers for B (row-indexed)
tiers_B = {}
for r in staff_rows:
    try:
        tiers_B[r] = int(ws2.cell(r, SD_B).value or 0)
    except ValueError:
        tiers_B[r] = 0  # non-numeric -> ineligible



def _cell_core_arr(r, c):
    core, arr = split_arrow(ws2.cell(r, c).value or "")
    return core.strip(), arr

for c in day_cols:
    # 1) Skip this day if a B already exists
    has_B = False
    for r in staff_rows:
        core, _ = _cell_core_arr(r, c)
        if core == "B":
            has_B = True
            break
    if has_B:
        continue

    # 2) Build tier -> candidate lists (A only, excluding PA (prev day PM))
    tier_to_candidates = {}
    for r in staff_rows:
        if tiers_B.get(r, 0) <= 0:
            continue  # ineligible
        core, _ = _cell_core_arr(r, c)
        if core != "A":
            continue

        # Exclude A if it's part of a P->A (PA) transition
        is_PA = False
        # If there is a previous day column
        if c != day_cols[0]:
            prev_col = day_cols[day_cols.index(c) - 1]
            prev_core, _ = _cell_core_arr(r, prev_col)
            if prev_core in PM_SET:
                is_PA = True
        # If it's the very first day, no previous day -> not PA
        if is_PA:
            continue  # forbidden per your rule

        tier_to_candidates.setdefault(tiers_B[r], []).append(r)

    if not tier_to_candidates:
        continue  # no eligible A anywhere → nothing to convert

    # 3) Process the lowest tier with at least one eligible A
    for tier in sorted(tier_to_candidates.keys()):
        cand_all = tier_to_candidates[tier]
        if not cand_all:
            continue  # empty tier → try next tier

        # Prefer A that starts A->Night (A–N–…)
        pref = []
        # If a next day exists
        if c != day_cols[-1]:
            next_col = day_cols[day_cols.index(c) + 1]
            for r in cand_all:
                nxt_core, _ = _cell_core_arr(r, next_col)
                if nxt_core in NIGHT_SET:
                    pref.append(r)

        chosen = random.choice(pref if pref else cand_all)
        # Convert chosen A -> B (preserve ↗)
        _, arr = _cell_core_arr(chosen, c)
        ws2.cell(chosen, c).value = "B" + arr

        # Done for this day (exactly one B/day; do NOT escalate tiers if we placed one)
        break
# === Special duty K (from SD+2), strict per-tier; forbid PA; prefer A→N→… ===
SD_K = SD + 2  # column for K tiers

# Read tiers for K (row-indexed)
tiers_K = {}
for r in staff_rows:
    try:
        tiers_K[r] = int(ws2.cell(r, SD_K).value or 0)
    except ValueError:
        tiers_K[r] = 0  # non-numeric -> ineligible



def _cell_core_arr(r, c):
    core, arr = split_arrow(ws2.cell(r, c).value or "")
    return core.strip(), arr

for c in day_cols:
    # 1) Skip if a K already exists on this day
    has_K = False
    for r in staff_rows:
        core, _ = _cell_core_arr(r, c)
        if core == "K":
            has_K = True
            break
    if has_K:
        continue

    # 2) Build tier -> candidate lists (A only, excluding PA (prev day PM))
    tier_to_candidates = {}
    # Locate this day's index once to avoid repeated lookups
    day_idx = day_cols.index(c)

    for r in staff_rows:
        if tiers_K.get(r, 0) <= 0:
            continue  # ineligible
        core, _ = _cell_core_arr(r, c)
        if core != "A":
            continue

        # Exclude A if it's part of a P->A transition
        is_PA = False
        if day_idx > 0:
            prev_col = day_cols[day_idx - 1]
            prev_core, _ = _cell_core_arr(r, prev_col)
            if prev_core in PM_SET:
                is_PA = True
        if is_PA:
            continue

        tier_to_candidates.setdefault(tiers_K[r], []).append(r)

    if not tier_to_candidates:
        continue  # no eligible A anywhere → nothing to convert

    # 3) Process the lowest tier that has any eligible A
    for tier in sorted(tier_to_candidates.keys()):
        cand_all = tier_to_candidates[tier]
        if not cand_all:
            continue

        # Prefer A that starts A->Night
        pref = []
        if day_idx < len(day_cols) - 1:
            next_col = day_cols[day_idx + 1]
            for r in cand_all:
                nxt_core, _ = _cell_core_arr(r, next_col)
                if nxt_core in NIGHT_SET:
                    pref.append(r)

        chosen = random.choice(pref if pref else cand_all)
        # Convert chosen A -> K (preserve ↗)
        _, arr = _cell_core_arr(chosen, c)
        ws2.cell(chosen, c).value = "K" + arr

        # Done for this day (one K/day; do NOT escalate to higher tiers)
        break




# === Special duty O® (from SD+12), weekends only ===
SD_ORES = SD + 12  # column for O® tiers

# Read tiers for O®
tiers_ORES = {}
for r in staff_rows:
    try:
        tiers_ORES[r] = int(ws2.cell(r, SD_ORES).value or 0)
    except ValueError:
        tiers_ORES[r] = 0  # ineligible if not numeric

# Helper: find week start and end indices given a day index
def get_week_bounds(day_idx):
    sun_indices = [i for i, cc in enumerate(day_cols)
                   if str(ws2.cell(4, cc).value or "").strip().upper() == "SUN"]
    week_start = 0
    for si in sun_indices:
        if si <= day_idx:
            week_start = si
        else:
            break
    week_end = (sun_indices[sun_indices.index(week_start)+1]
                if week_start in sun_indices and sun_indices.index(week_start)+1 < len(sun_indices)
                else len(day_cols))
    return week_start, week_end

# Helper: does staff have at least one ☆ in this week?
def has_star_in_week(r, day_idx):
    week_start, week_end = get_week_bounds(day_idx)
    for j in range(week_start, week_end):
        core, _ = split_arrow(ws2.cell(r, day_cols[j]).value or "")
        if core == "☆":
            return True
    return False

# Helper: does staff have another plain O in this week?
def has_another_O_in_week(r, day_idx):
    week_start, week_end = get_week_bounds(day_idx)
    count_O = 0
    for j in range(week_start, week_end):
        core, _ = split_arrow(ws2.cell(r, day_cols[j]).value or "")
        if core == "O":
            count_O += 1
    return count_O >= 2  # at least 2 O's including current

# Helper: does staff already have O® in this week?
def has_Ores_in_week(r, day_idx):
    week_start, week_end = get_week_bounds(day_idx)
    for j in range(week_start, week_end):
        val = str(ws2.cell(r, day_cols[j]).value or "")
        if val.startswith("O®"):
            return True
    return False

# Assign one O® per weekend day
for c in sorted(weekend):
    # skip if already has O® that day
    if any(str(ws2.cell(r, c).value or "").startswith("O®") for r in staff_rows):
        continue

    day_idx = day_cols.index(c)
    tier_to_candidates = {}
    for r in staff_rows:
        tier = tiers_ORES.get(r, 0)
        if tier <= 0:
            continue

        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if core != "O":
            continue
        if "↗" in str(ws2.cell(r, c).value or ""):  # exclude O↗
            continue

        # Condition: must have either a ☆ in week OR another O in week
        if not (has_star_in_week(r, day_idx) or has_another_O_in_week(r, day_idx)):
            continue

        # Condition: previous day not PM duty
        if day_idx > 0:
            prev_core, _ = split_arrow(ws2.cell(r, day_cols[day_idx-1]).value or "")
            if prev_core in PM_SET:
                continue

        # Passed all conditions
        tier_to_candidates.setdefault(tier, []).append((r, arr))

    if not tier_to_candidates:
        continue

    # Pick from lowest tier
    for tier in sorted(tier_to_candidates.keys()):
        cand = tier_to_candidates[tier]
        if not cand:
            continue

        # Prefer staff without O® in this week
        without_Ores = [(r, arr) for (r, arr) in cand if not has_Ores_in_week(r, day_idx)]
        if without_Ores:
            rsel, arrsel = random.choice(without_Ores)
        else:
            rsel, arrsel = random.choice(cand)

        ws2.cell(rsel, c).value = "O®" + arrsel
        break


# === Special duty WEⓦ (from SD+4), weekends only; exclude A in A–N–O; strict per-tier ===
SD_WE_EMW = SD + 4  # column for WEⓦ tiers

# Read tiers for WEⓦ (row-indexed)
tiers_WEW = {}
for r in staff_rows:
    try:
        tiers_WEW[r] = int(ws2.cell(r, SD_WE_EMW).value or 0)
    except ValueError:
        tiers_WEW[r] = 0  # non-numeric -> ineligible

def _core_arr(r, c):
    core, arr = split_arrow(ws2.cell(r, c).value or "")
    return core.strip(), arr

def _contains_w_mark(r, c):
    # any code already carrying the ⓦ marker
    val = str(ws2.cell(r, c).value or "")
    return "ⓦ" in val

# Iterate weekend/holiday columns only
for c in sorted(weekend):
    # 1) Skip if any ⓦ already exists this day
    if any(_contains_w_mark(r, c) for r in staff_rows):
        continue

    # 2) Build tier -> candidate doctor rows having A (and not A–N–O)
    day_idx = day_cols.index(c)

    def is_A_N_O(r):
        """True if this A at (r,c) starts an A–Night–Off pattern."""
        # need next 2 days to check; if out of bounds -> cannot be A–N–O
        if day_idx + 2 >= len(day_cols):
            return False
        c1, c2 = day_cols[day_idx + 1], day_cols[day_idx + 2]
        next_core, _ = _core_arr(r, c1)
        next2_core, _ = _core_arr(r, c2)
        return (next_core in NIGHT_SET) and (next2_core in OFF_SET)

    tier_to_A = {}
    for r in staff_rows:
        tier = tiers_WEW.get(r, 0)
        if tier <= 0:
            continue  # ineligible
        core, _ = _core_arr(r, c)
        if core != "A":
            continue
        if is_A_N_O(r):
            continue  # exclude A that is part of A–N–O
        tier_to_A.setdefault(tier, []).append(r)

    if not tier_to_A:
        continue  # nothing to do this day

    # 3) Use the lowest tier that has any eligible A; do NOT escalate beyond it
    for tier in sorted(tier_to_A.keys()):
        cand = tier_to_A[tier]
        if not cand:
            continue
        rsel = random.choice(cand)
        _, arr = _core_arr(rsel, c)
        ws2.cell(rsel, c).value = "A2ⓦ" + arr  # convert and preserve ↗
        break  # one assignment per day, stop here


# === Special duty E2ω (from SD+5), all days; strict per-tier ===
SD_E2w = SD + 5  # column for E2ω tiers

# Read tiers for E2ω
tiers_E2W = {}
for r in staff_rows:
    try:
        tiers_E2W[r] = int(ws2.cell(r, SD_E2w).value or 0)
    except ValueError:
        tiers_E2W[r] = 0  # non-numeric -> ineligible

def _contains_e2w(r, c):
    val = str(ws2.cell(r, c).value or "")
    return "E2ω" in val

# Loop over every day
for c in day_cols:
    # 1) Skip if already has E2ω
    if any(_contains_e2w(r, c) for r in staff_rows):
        continue

    # 2) Build tier → candidate staff who have P
    tier_to_P = {}
    for r in staff_rows:
        tier = tiers_E2W.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if core == "P":
            tier_to_P.setdefault(tier, []).append((r, arr))

    if not tier_to_P:
        continue  # no candidates today

    # 3) Use lowest tier that has candidates; do not escalate further
    for tier in sorted(tier_to_P.keys()):
        cand = tier_to_P[tier]
        if not cand:
            continue
        rsel, arrsel = random.choice(cand)
        ws2.cell(rsel, c).value = "E2ω" + arrsel  # assign, preserve ↗
        break  # only one per day


 # === Special duty ▽ (from SD+6), weekdays except WED; strict per-tier ===
SD_AEFU = SD + 6  # column for ▽ tiers

# Read tiers for ▽
tiers_V = {}
for r in staff_rows:
    try:
        tiers_V[r] = int(ws2.cell(r, SD_AEFU).value or 0)
    except ValueError:
        tiers_V[r] = 0  # non-numeric -> ineligible

def _contains_V(r, c):
    val = str(ws2.cell(r, c).value or "")
    return "A▽" in val

# Loop through all weekday columns (exclude weekends and WED)
for c in weekday:
    wd = str(ws2.cell(4, c).value or "").strip().upper()
    if wd == "WED":  # skip Wednesdays
        continue

    # 1) Skip if already has A▽
    if any(_contains_V(r, c) for r in staff_rows):
        continue

    # 2) Build tier → candidate staff who have A
    tier_to_A = {}
    for r in staff_rows:
        tier = tiers_V.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if core == "A":
            tier_to_A.setdefault(tier, []).append((r, arr))

    if not tier_to_A:
        continue  # no eligible A today

    # 3) Use lowest tier that has candidates; do not escalate further
    for tier in sorted(tier_to_A.keys()):
        cand = tier_to_A[tier]
        if not cand:
            continue
        rsel, arrsel = random.choice(cand)
        ws2.cell(rsel, c).value = "A▽" + arrsel  # assign, preserve ↗
        break  # only one per day



# === Special duty WDⓦ (from SD+9), weekdays only; convert Z2→Z2ⓦ ===
SD_WD_SeniorEMW = SD + 9  # column for WDⓦ tiers

# Read tiers for WDⓦ
tiers_WDX = {}
for r in staff_rows:
    try:
        tiers_WDX[r] = int(ws2.cell(r, SD_WD_SeniorEMW).value or 0)
    except ValueError:
        tiers_WDX[r] = 0  # invalid entry -> ineligible

def _contains_WDX(r, c):
    val = str(ws2.cell(r, c).value or "")
    return "ⓦ" in val  # already has ⓦ duty

# Loop through weekday cols
for c in weekday:
    # Skip if already has ⓦ
    if any(_contains_WDX(r, c) for r in staff_rows):
        continue

    # Build candidates by tier
    tier_to_Z2 = {}
    for r in staff_rows:
        tier = tiers_WDX.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if core == "Z2":
            tier_to_Z2.setdefault(tier, []).append((r, arr))

    # Assign if available
    if tier_to_Z2:
        for tier in sorted(tier_to_Z2.keys()):
            cand = tier_to_Z2[tier]
            if not cand:
                continue
            rsel, arrsel = random.choice(cand)
            ws2.cell(rsel, c).value = "Z2ⓦ" + arrsel
            break




# === Special duty ㊥ (from SD+7), on WED only; prefer Z→Z2㊥ else A→A2㊥ ===
SD_Toxi = SD + 7  # column for ㊥ tiers

# Read tiers for ㊥
tiers_M = {}
for r in staff_rows:
    try:
        tiers_M[r] = int(ws2.cell(r, SD_Toxi).value or 0)
    except ValueError:
        tiers_M[r] = 0  # non-numeric -> ineligible

def _contains_M(r, c):
    val = str(ws2.cell(r, c).value or "")
    return "㊥" in val  # already assigned

# Loop through all weekday columns, filter for Wednesday only
for c in weekday:
    wd = str(ws2.cell(4, c).value or "").strip().upper()
    if wd != "WED":
        continue

    # Skip if already has ㊥ duty
    if any(_contains_M(r, c) for r in staff_rows):
        continue

    # --- Step 1: Prefer Z candidates ---
    tier_to_Z = {}
    for r in staff_rows:
        tier = tiers_M.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if core == "Z2":
            tier_to_Z.setdefault(tier, []).append((r, arr))

    if tier_to_Z:
        for tier in sorted(tier_to_Z.keys()):
            cand = tier_to_Z[tier]
            if not cand:
                continue
            rsel, arrsel = random.choice(cand)
            ws2.cell(rsel, c).value = "Z2㊥" + arrsel
            break
        continue  # done for this day

    # --- Step 2: Fallback to A/A2 candidates ---
    tier_to_A = {}
    for r in staff_rows:
        tier = tiers_M.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if core in ("A", "A2"):   # allow both A and A2
            tier_to_A.setdefault(tier, []).append((r, core, arr))

    if tier_to_A:
        for tier in sorted(tier_to_A.keys()):
            cand = tier_to_A[tier]
            if not cand:
                continue
            rsel, core_sel, arrsel = random.choice(cand)
            ws2.cell(rsel, c).value = core_sel + "㊥" + arrsel
            break

# === Special duty ⓧ (from SD+8), on WED only; convert A→A2ⓧ ===
SD_tele = SD + 8  # column for ⓧ tiers

# Read tiers for ⓧ
tiers_X = {}
for r in staff_rows:
    try:
        tiers_X[r] = int(ws2.cell(r, SD_tele).value or 0)
    except ValueError:
        tiers_X[r] = 0  # invalid entry -> ineligible

def _contains_X(r, c):
    val = str(ws2.cell(r, c).value or "")
    return "ⓧ" in val  # already has ⓧ duty
# === Special duty ⓧ (from SD+8), on WED only; convert A/A2 -> Aⓧ or A2ⓧ ===
for c in weekday:
    wd = str(ws2.cell(4, c).value or "").strip().upper()
    if wd != "WED":
        continue

    # Skip if already has ⓧ duty
    if any(_contains_X(r, c) for r in staff_rows):
        continue

    # Build candidates by tier
    tier_to_A = {}
    for r in staff_rows:
        tier = tiers_X.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if core in ("A", "A2"):  # allow both A and A2
            tier_to_A.setdefault(tier, []).append((r, core, arr))

    # Assign if available
    if tier_to_A:
        for tier in sorted(tier_to_A.keys()):
            cand = tier_to_A[tier]
            if not cand:
                continue
            rsel, core_sel, arrsel = random.choice(cand)
            ws2.cell(rsel, c).value = core_sel + "ⓧ" + arrsel
            break

# === Special duty ♥ (from SD+10), every day ===
SD_resus = SD + 10  # column for ♥ tiers

# Read tiers for ♥
tiers_HEART = {}
for r in staff_rows:
    try:
        tiers_HEART[r] = int(ws2.cell(r, SD_resus).value or 0)
    except ValueError:
        tiers_HEART[r] = 0  # invalid entry -> ineligible

def _can_heart_APK(core):
    """Eligible for ♥ conversion on A/B/K"""
    return core in ("A", "B", "K")

def _can_heart_P(core):
    """Eligible for ♥ conversion on P"""
    return core == "P"

for c in day_cols:
    # --- Part 1: A/B/K -> ♥ ---
    tier_to_candidates = {}
    for r in staff_rows:
        tier = tiers_HEART.get(r, 0)
        if tier <= 0: 
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if _can_heart_APK(core):
            tier_to_candidates.setdefault(tier, []).append((r, core, arr))
    if tier_to_candidates:
        for tier in sorted(tier_to_candidates.keys()):
            cand = tier_to_candidates[tier]
            if not cand:
                continue
            rsel, core_sel, arrsel = random.choice(cand)
            ws2.cell(rsel, c).value = core_sel + "♥" + arrsel
            break

    # --- Part 2: P -> P♥ ---
    tier_to_candidates = {}
    for r in staff_rows:
        tier = tiers_HEART.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if _can_heart_P(core):
            tier_to_candidates.setdefault(tier, []).append((r, core, arr))
    if tier_to_candidates:
        for tier in sorted(tier_to_candidates.keys()):
            cand = tier_to_candidates[tier]
            if not cand:
                continue
            rsel, core_sel, arrsel = random.choice(cand)
            ws2.cell(rsel, c).value = core_sel + "♥" + arrsel
            break


    # === Special duty %¥ (from SD+11), every day ===
SD_walking_lab = SD + 11  # column for %¥ tiers

# Read tiers for %¥
tiers_PCTYEN = {}
for r in staff_rows:
    try:
        tiers_PCTYEN[r] = int(ws2.cell(r, SD_walking_lab).value or 0)
    except ValueError:
        tiers_PCTYEN[r] = 0  # invalid -> ineligible

def _can_pctyen_APK(core):
    # Only plain A/B/K are convertible here
    return core in ("A", "B", "K")

def _can_pctyen_P(core):
    return core in ("P", "E2", "S2")

for c in day_cols:
    # --- Part 1: A/B/K -> with %¥ ---
    tier_to_candidates = {}
    for r in staff_rows:
        tier = tiers_PCTYEN.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if _can_pctyen_APK(core):
            tier_to_candidates.setdefault(tier, []).append((r, core, arr))

    if tier_to_candidates:
        for tier in sorted(tier_to_candidates.keys()):
            cand = tier_to_candidates[tier]
            if not cand:
                continue
            rsel, core_sel, arrsel = random.choice(cand)
            ws2.cell(rsel, c).value = core_sel + "%¥" + arrsel
            break  # do exactly one A/B/K per day

    # --- Part 2: P -> P%¥ ---
    tier_to_candidates = {}
    for r in staff_rows:
        tier = tiers_PCTYEN.get(r, 0)
        if tier <= 0:
            continue
        core, arr = split_arrow(ws2.cell(r, c).value or "")
        if _can_pctyen_P(core):
            tier_to_candidates.setdefault(tier, []).append((r, core, arr))

    if tier_to_candidates:
        for tier in sorted(tier_to_candidates.keys()):
            cand = tier_to_candidates[tier]
            if not cand:
                continue
            rsel, core_sel, arrsel = random.choice(cand)
            ws2.cell(rsel, c).value = core_sel + "%¥" + arrsel
            break  # do exactly one P per day


    # 4) remaining A by CON1/CON2 → A2
    for r in staff_rows:
        if ranks[r].startswith(("CON1","CON2")):
            for c in day_cols:
                core, arr = split_arrow(ws2.cell(r,c).value)
                if core == "A":
                    ws2.cell(r,c).value = "A2" + arr

 
    # 5) Convert first 'O' -> 'RD' per staff per week (weeks start on SUN)
    # --- Per-week: if no star present, convert first 'O' -> '☆' ---
# assumes: day_cols (list of columns for this month), staff_rows, ws2, split_arrow()

# collect indices of all Sundays (in day_cols)
sun_idx = []
for idx, c in enumerate(day_cols):
    wd = str(ws2.cell(4, c).value or "").strip().upper()
    if wd == "SUN":
        sun_idx.append(idx)

if sun_idx:
    for k, start_idx in enumerate(sun_idx):
        # define the week as [this SUN, next SUN) or [this SUN, end)
        end_idx = sun_idx[k + 1] if (k + 1) < len(sun_idx) else len(day_cols)
        week_cols = day_cols[start_idx:end_idx]  # may be < 7 at the end

        for r in staff_rows:
            had_star = False
            # 1) If this week already contains a ☆, skip conversion
            for c in week_cols:
                core, _ = split_arrow(ws2.cell(r, c).value)
                if core.strip() == "☆":
                    had_star = True
                    break

            if had_star:
                continue

            # 2) No ☆ present → convert the first 'O' to '☆' (preserve ↗)
            converted = False
            for c in week_cols:
                core, arr = split_arrow(ws2.cell(r, c).value)
                if core.strip() == "O":
                    ws2.cell(r, c).value = "☆" + arr
                    converted = True
                    break  # only the first O per staff per week

            # 3) If still no ☆ after attempted conversion → print info
            if not converted:
                staff_name = str(ws2.cell(r, 4).value or f"Row{r}")
                week_start_date = ws2.cell(3, day_cols[start_idx]).value
                print(f"⚠️ {staff_name} has no ☆ in week starting {week_start_date}")
# else: no Sunday found -> per rules, skip conversion
 
 # --- First-week conversion (cross-month aware): if no ☆ in 7-day window, convert first O in current partial week to ☆ ---

def _core(v):
    return str(v or "").rstrip("↗").strip()

def _split_arrow(v):
    s = str(v or "")
    return (s[:-1], "↗") if s.endswith("↗") else (s, "")

# Find index positions (in day_cols) where weekday is SUN
sun_idx = [i for i, c in enumerate(day_cols)
           if str(ws2.cell(4, c).value or "").strip().upper() == "SUN"]

# Only act if the first week is incomplete (i.e., first Sunday is not at index 0)
if sun_idx and sun_idx[0] > 0:
    # This month's partial first week columns: from month start up to day before first Sunday
    first_week_cols = day_cols[:sun_idx[0]]

    # Build previous-month tail columns (chronological), up to 7-len(first_week_cols)
    # We look left of START, but only keep columns that look like days (row 3 has a value)
    prev_cols = []
    c = START - 1
    while c >= 1 and len(prev_cols) < 7:
        if ws2.cell(3, c).value is not None:
            prev_cols.append(c)
        c -= 1
    prev_cols = list(reversed(prev_cols))  # oldest -> latest

    prev_needed = max(0, 7 - len(first_week_cols))
    prev_window = prev_cols[-prev_needed:] if prev_needed > 0 and len(prev_cols) >= prev_needed else []

    # Helper: any star in these absolute columns?
    def _has_star(r, cols):
        for cc in cols:
            if _core(ws2.cell(r, cc).value) == "☆":
                return True
        return False

    # For each staff row, apply the rule
    for r in staff_rows:
        # If any ☆ exists in the combined 7-day window (last month tail + current partial week), skip conversion
        if (prev_window and _has_star(r, prev_window)) or _has_star(r, first_week_cols):
            continue

        # Otherwise, convert the first 'O' in THIS MONTH'S partial week to '☆' (preserve ↗)
        for c in first_week_cols:
            core, arr = _split_arrow(ws2.cell(r, c).value)
            if core.strip() == "O":
                ws2.cell(r, c).value = "☆" + arr
                break

    # 6) Bulk conversions (preserve ↗ and any special duty marks)
    #    - All P* → E2* ; all other P → E2
    #    - All A → A2 IF tomorrow is not a night (N*, N, N3, N.). Last day counts as "not a night".
    NIGHT_SET = {"N*", "N", "N3", "N."}

    for r in staff_rows:
        for idx, c in enumerate(day_cols):
            raw_val = str(ws2.cell(r, c).value or "")
            core, arr = split_arrow(raw_val)

            # --- Skip already-converted forms
            if core.startswith(("E2", "A2")):
                continue

            # --- Convert P → E2 (keep suffix & arrow)
            if core.startswith("P"):
                if core == "P*":
                    ws2.cell(r, c).value = "E2*" + arr
                else:
                    # preserve any extra suffix after P (e.g. P♥ -> E2♥)
                    ws2.cell(r, c).value = "E2" + core[1:] + arr
                continue

            # --- Convert A → A2 (skip "AL", "A2*", etc.)
            if core.startswith("A") and not core.startswith("AL"):
                if idx == len(day_cols) - 1:
                    # last day: safe
                    ws2.cell(r, c).value = "A2" + core[1:] + arr
                else:
                    next_core, _ = split_arrow(ws2.cell(r, day_cols[idx + 1]).value or "")
                    if next_core not in NIGHT_SET:
                        ws2.cell(r, c).value = "A2" + core[1:] + arr
    wb2.save("Roster_Output2.xlsx")
    print("✅ Written Roster_Output2.xlsx")



if toggle_val < 3:
    exit(0)   # clean exit

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SRC_FILE = "Roster_Output2.xlsx"   # source of duties
SRC_SHEET = "Sheet1"

TGT_FILE = "v0.xlsx"   # template to rewrite

wb_tgt = load_workbook(TGT_FILE)

# always pick the first (default) sheet
TGT_SHEET = wb_tgt.sheetnames[0]
ws_tgt = wb_tgt[TGT_SHEET]
OUT_FILE = "Roster_Output3.xlsx"          # output file

# ---------- helpers ----------
def norm_text(v: object) -> str:
    s = str(v or "")
    return " ".join(s.split()).upper()

def find_anchor(ws, needle="NAME OF", max_rows=30, max_cols=100):
    nd = needle.upper()
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            txt = norm_text(ws.cell(r, c).value)
            if nd in txt:
                return r, c
    raise RuntimeError(f'Could not find anchor containing "{needle}" '
                       f'(searched {max_rows}×{max_cols}).')

def find_start_col(ws):
    c = 1
    while c <= ws.max_column:
        if norm_text(ws.cell(1, c).value) == "START":
            return c + 1
        c += 1
    raise RuntimeError('Could not find "START" in row 1.')

def list_day_cols(ws, start_col):
    cols = []
    c = start_col
    while ws.cell(3, c).value is not None:
        cols.append(c)
        c += 1
    if not cols:
        raise RuntimeError("No day columns found (row 3 empty at start).")
    return cols

def iter_staff_rows(ws, row_start=6, name_col=4):
    r = row_start
    while True:
        name = ws.cell(r, name_col).value
        if not name:
            break
        yield r
        r += 1

def row_hidden(ws, r):
    dim = ws.row_dimensions.get(r)
    return bool(dim and dim.hidden)


# ---------- Check D4 toggle first ----------
wb_src = load_workbook(SRC_FILE)
ws_src = wb_src[SRC_SHEET]


# ---------- If toggle = 3, proceed ----------
print("➡️  Proceeding with write_excel block...")

# 1) Build {doctor -> [duties]}
src_start_col = find_start_col(ws_src)
src_day_cols = list_day_cols(ws_src, src_start_col)

SRC_NAME_COL = 4
SRC_ROW_START = 6

duties_by_doctor = {}
for r in iter_staff_rows(ws_src, SRC_ROW_START, SRC_NAME_COL):
    name = str(ws_src.cell(r, SRC_NAME_COL).value or "").strip()
    if not name:
        continue
    row_duties = [str(ws_src.cell(r, c).value or "").strip() for c in src_day_cols]
    duties_by_doctor[name] = row_duties

num_days = len(src_day_cols)
#print(f"Source days detected: {num_days}; doctors: {len(duties_by_doctor)}")

# 2) Write into target workbook
wb_tgt = load_workbook(TGT_FILE)
ws_tgt = wb_tgt[TGT_SHEET]

anchor_r, anchor_c = find_anchor(ws_tgt, "Name of")
name_col = anchor_c + 1
first_day_col = anchor_c + 2
first_doc_row = anchor_r + 3

#print(f'Anchor at {get_column_letter(anchor_c)}{anchor_r}; '
#      f'Name col = {get_column_letter(name_col)}, '
#      f'First day col = {get_column_letter(first_day_col)}, '
#      f'First doc row = {first_doc_row}')

written_rows = 0
skipped_locked = 0
skipped_hidden = 0
skipped_no_match = 0

r = first_doc_row
while r <= ws_tgt.max_row:
    raw_name = ws_tgt.cell(r, name_col).value
    if raw_name is None or str(raw_name).strip() == "":
        r += 1
        continue

    if row_hidden(ws_tgt, r):
        skipped_hidden += 1
        r += 1
        continue

    doc_name = str(raw_name).strip()
    duties = duties_by_doctor.get(doc_name)
    if not duties:
        skipped_no_match += 1
        r += 1
        continue

    for d in range(num_days):
        c = first_day_col + d
        cell = ws_tgt.cell(r, c)
        if cell.protection.locked:
            skipped_locked += 1
            continue
        cell.value = duties[d]
    written_rows += 1
    r += 1

wb_tgt.save(OUT_FILE)
print(f"✅ Wrote duties to '{OUT_FILE}'")
#print(f"Summary: rows_written={written_rows}, skipped_hidden_rows={skipped_hidden}, "
#      f"skipped_no_name_match={skipped_no_match}, skipped_locked_cells={skipped_locked}")
